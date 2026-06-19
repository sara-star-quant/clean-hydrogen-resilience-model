"""Grant packaging: bundle a scheme-specific archive ready for submission.

Takes the generic project (report, grant template, references, charts) and assembles a
dated zip per scheme. CSV-fallback budget if openpyxl is not installed.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import json
import zipfile
from pathlib import Path
from typing import Any

from .paths import REPO_ROOT
from .tech import load_params

__all__ = ["PACKAGE_MANIFESTS", "available_schemes", "package_grant"]


COMMON: tuple[str, ...] = (
    "DISCLAIMER.md",
    "report/00_executive_summary.md",
    "report/references.md",
    "report/_generated_tables.md",
)

DISTRICT_FOCUSED: tuple[str, ...] = (
    *COMMON,
    "report/01_problem_and_scope.md",
    "report/02_technology_landscape.md",
    "report/04_district_track.md",
    "report/05_economics_lcoe_tco.md",
    "report/06_environmental_lca.md",
    "report/07_risk_register.md",
    "report/08_roadmap_milestones.md",
    "report/10_supply_resilience_and_regions.md",
    "grants/_shared_narrative.md",
)

CAR_FOCUSED: tuple[str, ...] = (
    *COMMON,
    "report/01_problem_and_scope.md",
    "report/02_technology_landscape.md",
    "report/03_car_track.md",
    "report/05_economics_lcoe_tco.md",
    "report/07_risk_register.md",
    "report/08_roadmap_milestones.md",
    "grants/_shared_narrative.md",
)

PACKAGE_MANIFESTS: dict[str, tuple[str, ...]] = {
    "horizon_europe":     (*DISTRICT_FOCUSED, "grants/eu_horizon_europe_cluster5.md"),
    "eu_innovation_fund": (*DISTRICT_FOCUSED, "grants/eu_innovation_fund.md"),
    "arpa_e":             (*CAR_FOCUSED, "grants/us_arpa_e.md"),
    "doe_h2_programs":    (*DISTRICT_FOCUSED, "grants/us_doe_h2_programs.md"),
    "arena":              (*DISTRICT_FOCUSED, "grants/au_arena.md"),
    "eeca_callaghan":     (*DISTRICT_FOCUSED, "grants/nz_eeca_callaghan.md"),
}

# Per-scheme cost categories for the budget table. Each entry is a list of
# (category_label, default_share_of_total).
BUDGET_CATEGORIES: dict[str, list[tuple[str, float]]] = {
    "horizon_europe": [
        ("Personnel", 0.45),
        ("Subcontracting", 0.10),
        ("Equipment", 0.20),
        ("Other direct", 0.05),
        ("Indirect (25 percent flat)", 0.20),
    ],
    "eu_innovation_fund": [
        ("Personnel", 0.30),
        ("Subcontracting", 0.10),
        ("Equipment and works", 0.45),
        ("Other direct", 0.05),
        ("Indirect", 0.10),
    ],
    "arpa_e": [
        ("Direct labour", 0.45),
        ("Fringe", 0.12),
        ("Travel", 0.03),
        ("Materials and supplies", 0.08),
        ("Equipment", 0.10),
        ("Subcontracts", 0.07),
        ("Indirect (negotiated rate placeholder 50 percent)", 0.15),
    ],
    "doe_h2_programs": [
        ("Direct", 0.80),
        ("Indirect", 0.20),
    ],
    "arena": [
        ("Milestone 1 (design and permit)", 0.10),
        ("Milestone 2 (procurement)", 0.20),
        ("Milestone 3 (construction)", 0.40),
        ("Milestone 4 (commissioning)", 0.20),
        ("Milestone 5 (measurement)", 0.10),
    ],
    "eeca_callaghan": [
        ("Direct R&D", 0.80),
        ("Indirect", 0.20),
    ],
}

DEFAULT_TOTAL_USD = {
    "horizon_europe": 30_000_000,
    "eu_innovation_fund": 25_000_000,
    "arpa_e": 1_000_000,
    "doe_h2_programs": 36_000_000,
    "arena": 24_000_000,
    "eeca_callaghan": 1_500_000,
}


def available_schemes() -> list[str]:
    """Return the list of grant scheme names supported by the packager."""
    return sorted(PACKAGE_MANIFESTS)


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _budget_csv(scheme: str, total_usd: float) -> bytes:
    """Render the scheme's budget table as CSV bytes."""
    cats = BUDGET_CATEGORIES[scheme]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["category", "share_of_total", "amount_usd"])
    for label, share in cats:
        writer.writerow([label, f"{share:.2f}", f"{total_usd * share:,.0f}"])
    writer.writerow(["TOTAL", "1.00", f"{total_usd:,.0f}"])
    writer.writerow([])
    writer.writerow(["Indicative budget; numbers are scenario outputs derived from cited public sources. Replace before submission. See DISCLAIMER.md."])
    return buf.getvalue().encode("utf-8")


def _budget_xlsx_or_csv(scheme: str, total_usd: float) -> tuple[str, bytes]:
    """Render the budget. XLSX if openpyxl is available, otherwise CSV.

    Returns (filename_in_archive, bytes).
    """
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
    except ImportError:
        return f"budget_{scheme}.csv", _budget_csv(scheme, total_usd)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        # Workbook() always seeds an active sheet; this branch is for the type checker.
        ws = wb.create_sheet("Budget")
    ws.title = "Budget"
    ws.append(["category", "share_of_total", "amount_usd"])
    for label, share in BUDGET_CATEGORIES[scheme]:
        ws.append([label, share, round(total_usd * share)])
    ws.append(["TOTAL", 1.0, round(total_usd)])
    ws.append([])
    ws.append(["Indicative budget; numbers are scenario outputs derived from cited public sources. Replace before submission. See DISCLAIMER.md."])
    buf = io.BytesIO()
    wb.save(buf)
    return f"budget_{scheme}.xlsx", buf.getvalue()


def package_grant(
    scheme: str,
    out_dir: Path,
    *,
    repo_root: Path | None = None,
    today: dt.date | None = None,
    total_usd: float | None = None,
) -> Path:
    """Bundle a scheme-specific submission package and write it to disk.

    Returns the path to the created zip archive.
    """
    if scheme not in PACKAGE_MANIFESTS:
        raise ValueError(
            f"unknown scheme {scheme!r}. Known: {sorted(PACKAGE_MANIFESTS)}"
        )
    repo_root = repo_root or REPO_ROOT
    today = today or dt.date.today()
    total_usd = total_usd if total_usd is not None else float(DEFAULT_TOTAL_USD[scheme])
    manifest_files = PACKAGE_MANIFESTS[scheme]

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    archive_name = f"electicity_{scheme}_{today.isoformat()}.zip"
    archive_path = out_dir / archive_name

    manifest_records: list[dict[str, Any]] = []

    # Validate all expected files exist before opening the zip.
    missing = [f for f in manifest_files if not (repo_root / f).is_file()]
    if missing:
        raise FileNotFoundError(
            f"manifest references missing files: {missing}"
        )

    # Touch load_params so a malformed tech_params.yaml fails before the archive lands.
    load_params()

    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for rel in manifest_files:
            data = (repo_root / rel).read_bytes()
            zf.writestr(rel, data)
            manifest_records.append({
                "path": rel,
                "size_bytes": len(data),
                "sha256": _sha256_bytes(data),
            })

        budget_name, budget_bytes = _budget_xlsx_or_csv(scheme, total_usd)
        zf.writestr(budget_name, budget_bytes)
        manifest_records.append({
                "path": budget_name,
                "size_bytes": len(budget_bytes),
                "sha256": _sha256_bytes(budget_bytes),
            })

        manifest_payload = {
            "scheme": scheme,
            "date": today.isoformat(),
            "total_usd": total_usd,
            "files": manifest_records,
            "note": (
                "Indicative submission package. Numbers are scenario outputs "
                "derived from cited public sources. Replace placeholders and "
                "the budget total before submission. See DISCLAIMER.md inside "
                "the archive."
            ),
        }
        manifest_bytes = json.dumps(manifest_payload, indent=2).encode("utf-8")
        zf.writestr("manifest.json", manifest_bytes)

    return archive_path
